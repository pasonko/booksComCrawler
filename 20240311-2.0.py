from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
from bs4 import BeautifulSoup
import time
import re
import openpyxl
from pathlib import Path
import pandas as pd
from webdriver_manager.chrome import ChromeDriverManager  # 导入 ChromeDriverManager

# 设置搜尋引擎的網址
url = 'https://www.books.com.tw/?loc=tw_logo_001'

# 打開桌面上的 excel 檔案，並取得 A column 中所有的關鍵字
desktop = Path(r".\\")
workbook = openpyxl.load_workbook(desktop.joinpath("search.xlsx"))
worksheet = workbook.active
keywords = []
for row in worksheet.iter_rows(min_row=1, max_col=1, values_only=True):
    keyword = row[0]  # 因為我們只取 A column，所以這裡只需要索引為 0 的元素
    keywords.append(keyword)

new_file_name = "result.xlsx"
new_file_path = ".\\" + new_file_name
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# 將欄位標題寫入新的 excel 檔案
new_sheet['A1'] = "關鍵字"
new_sheet['B1'] = "書名"
new_sheet['C1'] = "作者"
new_sheet['D1'] = "出版社"
new_sheet['E1'] = "ISBN"


#options.add_argument('--headless')  # 在后台运行 Chrome，不显示界面

# driver_path = r"C:\Program Files\Google\Chrome\Application\chromedriver.exe"
# service = Service(driver_path)
# options = webdriver.ChromeOptions()
# browser = webdriver.Chrome(service=service, options=options)


# 使用 Chrome 瀏覽器開啟網頁，進行搜尋
options = webdriver.ChromeOptions()
# 使用 Service 对象来启动 ChromeDriver
service = Service(ChromeDriverManager().install())
browser = webdriver.Chrome(service=service, options=options)



for i, keyword in enumerate(keywords):
    
    isbn = None
    author = None
    translator = None
    publisher = None

    if keyword is None:
        break
    else:
        # 打開搜尋引擎，輸入關鍵字並按下 Enter 鍵
        browser.get(url)
        search_bar = browser.find_element(By.ID, 'key')
        search_bar.send_keys(keyword)
        search_bar.send_keys(Keys.ENTER)

        wait = WebDriverWait(browser, 10)

        try:
            # XPATH for the first search result (modify if needed)
            my_result = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@class="table-searchbox clearfix"]//div[@class="mod2 table-container"]//div[@class="table-tr"]//div[@class="table-td"]//a')))
            my_result_url = my_result.get_attribute("href")  # Extract URL from the link

            # Print the extracted URL
            print(f"Extracted URL for {keyword}: {my_result_url}")

        except:
            print(f"Could not find result for keyword: {keyword}")
            row = i + 2
            new_sheet['A' + str(row)] = keyword
            new_sheet['B' + str(row)] = "查無資料"
            continue


        # 進入我要的搜尋結果的網頁
        browser.get(my_result_url)
        try:
            element = WebDriverWait(browser, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'bd'))
            )
            # 元素已經出現，可以繼續下一步操作
        except:
            # 等待超時，執行例外處理程式碼
            print("等待超時，找不到元素")

        # 取得書籍資訊
        try:
            element = browser.find_element(By.CLASS_NAME, 'mod_b.type02_m058.clearfix')
        except:
            element = None

        if element:
            # 執行後續程式碼
            a = browser.find_element(By.CLASS_NAME, 'type02_p003')
            test = a.text

            b = browser.find_element(By.CLASS_NAME, 'mod_b.type02_m058.clearfix')
            test2 = b.text

            book = browser.find_element(By.CLASS_NAME, 'mod.type02_p002.clearfix')
            bookname = book.text

            if "(電子書)" in bookname:
                bookname = bookname.replace("(電子書)", "")

            isbn = re.findall(r'ISBN：\s*(.+?)\s*\n', test2)
            author = re.findall(r'作者：\s*(.+?)\s*\n', test)
            translator = re.findall(r'譯者：\s*(.+?)\s*\n', test)
            publisher = re.findall(r'出版社：\s*(.+?)\s*\n', test)

            print(keyword)
            print(bookname)
            if len(isbn) > 0:
                print('ISBN：', isbn[0])
            else:
                print('ISBN：無')
            if len(author) > 0:
                print('作者：', author[0])
            else:
                print('作者：無')

            if len(translator) > 0:
                print('譯者：', translator[0])
            else:
                print('譯者：無')

            if len(publisher) > 0:
                print('出版社：', publisher[0])
            else:
                print('出版社：無')

            # 將搜尋結果寫入新的 excel 檔案
            row = i + 2
            new_sheet['A' + str(row)] = keyword
            new_sheet['B' + str(row)] = bookname if bookname else "無"
            if translator:
                new_sheet['C' + str(row)] = "{} / {}".format(author[0], translator[0])
            else:
                if len(author) > 0:
                    new_sheet['C' + str(row)] = author[0]
            new_sheet['E' + str(row)] = isbn[0] if isbn else "無"
            new_sheet['D' + str(row)] = publisher[0] if publisher else "無"
            new_wb.save(new_file_path)
        else:
            # 處理找不到元素的情況
            row = i + 2
            new_sheet['A' + str(row)] = keyword
            new_sheet['B' + str(row)] = "無"
            new_sheet['C' + str(row)] = "無"
            new_sheet['E' + str(row)] = "無"
            new_sheet['D' + str(row)] = "無"
            new_wb.save(new_file_path)

new_wb.save(new_file_path)
# 關閉瀏覽器
browser.quit()